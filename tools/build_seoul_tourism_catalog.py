#!/usr/bin/env python3
"""서울시 공식 XLSX+SHP를 Earthus 121장소 중심점 카탈로그로 변환한다.

좌표를 지명으로 추정하지 않는다. SHP polygon의 면적 가중 중심점을 계산하고,
같은 배포 묶음의 DBF AREA_CD를 XLSX 코드·영문명과 결합한다.
"""

from __future__ import annotations

import argparse
import json
import struct
from pathlib import Path

from openpyxl import load_workbook


def read_xlsx(path: Path) -> dict[str, dict]:
    sheet = load_workbook(path, data_only=True, read_only=True).active
    places: dict[str, dict] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = str(row[2] or "").strip()
        if not code:
            continue
        places[code] = {
            "code": code,
            "category": str(row[0] or "").strip(),
            "nameKo": str(row[3] or "").strip(),
            "nameEn": str(row[4] or "").strip(),
        }
    return places


def read_dbf(path: Path) -> list[dict[str, str]]:
    data = path.read_bytes()
    count = struct.unpack_from("<I", data, 4)[0]
    header_len, record_len = struct.unpack_from("<HH", data, 8)
    fields = []
    pos = 32
    while pos < header_len and data[pos] != 0x0D:
        desc = data[pos : pos + 32]
        name = desc[:11].split(b"\0", 1)[0].decode("ascii")
        fields.append((name, desc[16]))
        pos += 32
    rows = []
    for index in range(count):
        record = data[header_len + index * record_len : header_len + (index + 1) * record_len]
        if not record or record[0:1] == b"*":
            continue
        cursor = 1
        row: dict[str, str] = {}
        for name, length in fields:
            raw = record[cursor : cursor + length]
            cursor += length
            row[name] = raw.decode("utf-8", "replace").strip().strip("\0")
        rows.append(row)
    return rows


def ring_centroid(points: list[tuple[float, float]]) -> tuple[float, float, float]:
    """signed twice-area와 centroid numerator를 반환한다."""
    if len(points) < 3:
        return 0.0, 0.0, 0.0
    cross_sum = cx_sum = cy_sum = 0.0
    for a, b in zip(points, points[1:] + points[:1]):
        cross = a[0] * b[1] - b[0] * a[1]
        cross_sum += cross
        cx_sum += (a[0] + b[0]) * cross
        cy_sum += (a[1] + b[1]) * cross
    if abs(cross_sum) < 1e-16:
        return 0.0, 0.0, 0.0
    return cross_sum, cx_sum / (3 * cross_sum), cy_sum / (3 * cross_sum)


def polygon_centroid(parts: list[list[tuple[float, float]]]) -> tuple[float, float]:
    weighted_x = weighted_y = weight = 0.0
    for points in parts:
        signed_area2, cx, cy = ring_centroid(points)
        weighted_x += cx * signed_area2
        weighted_y += cy * signed_area2
        weight += signed_area2
    if abs(weight) > 1e-16:
        return weighted_x / weight, weighted_y / weight
    flat = [point for part in parts for point in part]
    return sum(p[0] for p in flat) / len(flat), sum(p[1] for p in flat) / len(flat)


def read_shp(path: Path) -> list[dict]:
    data = path.read_bytes()
    if len(data) < 100 or struct.unpack_from(">I", data, 0)[0] != 9994:
        raise ValueError("invalid shapefile header")
    records = []
    pos = 100
    while pos + 8 <= len(data):
        _, words = struct.unpack_from(">II", data, pos)
        pos += 8
        body = data[pos : pos + words * 2]
        pos += words * 2
        if len(body) < 44:
            continue
        shape_type = struct.unpack_from("<I", body, 0)[0]
        if shape_type not in (5, 15, 25):
            raise ValueError(f"unsupported shape type {shape_type}")
        xmin, ymin, xmax, ymax = struct.unpack_from("<4d", body, 4)
        part_count, point_count = struct.unpack_from("<2I", body, 36)
        indexes = list(struct.unpack_from(f"<{part_count}I", body, 44))
        point_offset = 44 + 4 * part_count
        points = [struct.unpack_from("<2d", body, point_offset + i * 16) for i in range(point_count)]
        ends = indexes[1:] + [point_count]
        parts = [points[start:end] for start, end in zip(indexes, ends)]
        lon, lat = polygon_centroid(parts)
        records.append({
            "lat": round(lat, 7), "lon": round(lon, 7),
            "bbox": [round(xmin, 7), round(ymin, 7), round(xmax, 7), round(ymax, 7)],
        })
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--shape-dir", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()

    names = read_xlsx(args.xlsx)
    shp_files = list(args.shape_dir.glob("*.shp"))
    dbf_files = list(args.shape_dir.glob("*.dbf"))
    if len(shp_files) != 1 or len(dbf_files) != 1:
        raise ValueError(f"expected one SHP and DBF in {args.shape_dir}")
    attrs = read_dbf(dbf_files[0])
    shapes = read_shp(shp_files[0])
    if not (len(names) == len(attrs) == len(shapes) == 121):
        raise ValueError(f"expected 121 records, got xlsx={len(names)} dbf={len(attrs)} shp={len(shapes)}")

    places = []
    for attr, shape in zip(attrs, shapes):
        code = attr.get("AREA_CD", "")
        if code not in names:
            raise ValueError(f"AREA_CD missing from XLSX: {code}")
        row = {**names[code], **shape, "geometrySource": "서울시 주요 121장소 영역"}
        if attr.get("AREA_NM") and attr["AREA_NM"] != row["nameKo"]:
            raise ValueError(f"name mismatch {code}: {attr['AREA_NM']} != {row['nameKo']}")
        places.append(row)
    places.sort(key=lambda row: row["code"])

    output = {
        "schemaVersion": "earthus.tourism-place-catalog.v1",
        "generatedFrom": "서울시 2026-04-02 공식 배포 XLSX·SHP",
        "source": {
            "publisher": "서울특별시",
            "dataset": "서울시 실시간 도시데이터",
            "datasetUrl": "https://data.seoul.go.kr/dataList/OA-21285/F/1/datasetView.do",
            "files": ["서울시 주요 121장소 목록.xlsx", "서울시 주요 121장소 영역.zip"],
            "license": "공공누리 제1유형",
            "redisplay": "출처표시 · 상업적 이용 및 변경 가능",
        },
        "places": places,
    }
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {len(places)} places to {args.out}")


if __name__ == "__main__":
    main()
